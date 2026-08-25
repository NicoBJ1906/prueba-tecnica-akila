"""Tests del pipeline completo: orden de etapas, idempotencia y salida a Excel."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from triaje.estado import RegistroProcesados
from triaje.excel import COLUMNAS_EXIGIDAS, escribir
from triaje.modelos import ESTADO_REVISION, ESTADO_SIN_ACCION
from triaje.pipeline import ejecutar, leer_correos

RAIZ = Path(__file__).resolve().parents[2]
CSV_CORREOS = RAIZ / "data" / "correos_clientes.csv"


class TestLectura:
    def test_lee_los_quince_correos_de_la_muestra(self):
        assert len(leer_correos(CSV_CORREOS)) == 15

    def test_soporta_el_bom_de_excel(self):
        """El CSV entregado viene con BOM: la primera columna debe leerse igual."""
        correos = leer_correos(CSV_CORREOS)
        assert correos[0].remitente == "maria.lopez@gmail.com"
        assert correos[0].fecha_recepcion.hour == 8


class TestOrdenDeEtapas:
    def test_los_duplicados_del_dia_se_descartan(self, config):
        correos = leer_correos(CSV_CORREOS)
        resultado = ejecutar(correos, config)
        motivos = [d.regla for d in resultado.descartes]
        assert "deduplicación" in motivos

    def test_las_notificaciones_automaticas_no_llegan_al_seguimiento(self, config):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        remitentes = {f.remitente for f in resultado.filas}
        assert "notificaciones@bancolombia.com.co" not in remitentes

    def test_la_cortesia_no_genera_tarea_ni_responsable(self, config):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        cortesias = [f for f in resultado.filas if f.estado == ESTADO_SIN_ACCION]
        assert cortesias
        assert all(f.responsable == "—" for f in cortesias)
        assert all("Ninguna" in f.accion for f in cortesias)

    def test_el_desistimiento_va_a_revision_humana(self, config):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        jorge = [f for f in resultado.filas if "jorge" in f.remitente]
        assert jorge and all(f.estado == ESTADO_REVISION for f in jorge)

    def test_un_desistimiento_nunca_queda_en_urgencia_baja(self, config):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        jorge = [f for f in resultado.filas if "jorge" in f.remitente]
        assert all(f.urgencia in {"Media", "Alta"} for f in jorge)

    def test_el_correo_sin_contexto_va_a_revision(self, config):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        carlos = [f for f in resultado.filas if "carlos.medina" in f.remitente]
        assert carlos and all(f.estado == ESTADO_REVISION for f in carlos)

    def test_toda_fila_tiene_fecha_cliente_y_responsable(self, config):
        """El enunciado señala que hoy se olvidan campos: aquí no puede pasar."""
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        for fila in resultado.filas:
            assert fila.fecha is not None
            assert fila.cliente.strip()
            assert fila.responsable.strip()
            assert fila.accion.strip()


class TestIdempotencia:
    def test_ejecutar_dos_veces_no_duplica_filas(self, config, tmp_path):
        correos = leer_correos(CSV_CORREOS)
        registro = RegistroProcesados(tmp_path / "estado.json")

        primera = ejecutar(correos, config, registro_procesados=registro)
        registro.guardar()

        segunda = ejecutar(correos, config, registro_procesados=RegistroProcesados(tmp_path / "estado.json"))

        assert primera.filas
        assert segunda.filas == []
        assert len(segunda.ya_procesados) == len(correos)

    def test_un_correo_nuevo_si_se_procesa(self, config, tmp_path, correo):
        registro = RegistroProcesados(tmp_path / "estado.json")
        ejecutar(leer_correos(CSV_CORREOS), config, registro_procesados=registro)
        registro.guardar()

        nuevo = correo(cuerpo="Consulta sobre la entrega del apartamento 1801",
                       fecha="2026-07-21 09:00")
        segunda = ejecutar([nuevo], config, registro_procesados=RegistroProcesados(tmp_path / "estado.json"))
        assert len(segunda.filas) == 1

    def test_el_mismo_texto_otro_dia_no_es_un_duplicado(self, config, correo):
        """Reiterar una semana después es una petición nueva, no un duplicado."""
        texto = "¿Cuándo entregan el apartamento 803?"
        correos = [correo(cuerpo=texto, fecha="2026-07-20 09:00"),
                   correo(cuerpo=texto, fecha="2026-07-27 09:00")]
        resultado = ejecutar(correos, config)
        assert len(resultado.filas) == 2
        assert not resultado.descartes

    def test_un_registro_corrupto_no_detiene_el_proceso(self, config, tmp_path):
        ruta = tmp_path / "estado.json"
        ruta.write_text("{ esto no es json", encoding="utf-8")
        registro = RegistroProcesados(ruta)
        assert len(registro) == 0
        resultado = ejecutar(leer_correos(CSV_CORREOS), config, registro_procesados=registro)
        assert resultado.filas


class TestExcel:
    def test_las_seis_primeras_columnas_son_las_del_enunciado(self, config, tmp_path):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        ruta = escribir(resultado, tmp_path / "seguimiento.xlsx")

        hoja = load_workbook(ruta)["Seguimiento"]
        cabecera = [celda.value for celda in hoja[1]]
        assert cabecera[:6] == list(COLUMNAS_EXIGIDAS)
        assert cabecera[:6] == ["Fecha", "Cliente", "Tipo", "Urgencia", "Acción", "Responsable"]

    def test_el_libro_tiene_las_cuatro_hojas(self, config, tmp_path):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        libro = load_workbook(escribir(resultado, tmp_path / "seguimiento.xlsx"))
        assert set(libro.sheetnames) == {"Resumen", "Seguimiento", "Para revisar", "Descartados"}

    def test_escribe_una_fila_por_asunto(self, config, tmp_path):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        hoja = load_workbook(escribir(resultado, tmp_path / "seguimiento.xlsx"))["Seguimiento"]
        assert hoja.max_row == len(resultado.filas) + 1

    def test_la_hoja_de_revision_recoge_solo_lo_pendiente(self, config, tmp_path):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        libro = load_workbook(escribir(resultado, tmp_path / "seguimiento.xlsx"))
        assert libro["Para revisar"].max_row == len(resultado.para_revision) + 1

    def test_los_descartes_quedan_registrados_con_su_motivo(self, config, tmp_path):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        hoja = load_workbook(escribir(resultado, tmp_path / "seguimiento.xlsx"))["Descartados"]
        motivos = [fila[3] for fila in hoja.iter_rows(min_row=2, values_only=True)]
        assert len(motivos) == len(resultado.descartes)
        assert all(m for m in motivos)

    def test_volver_a_escribir_amplia_el_libro_sin_perder_lo_anterior(self, config, tmp_path):
        ruta = tmp_path / "seguimiento.xlsx"
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)

        escribir(resultado, ruta)
        filas_tras_la_primera = load_workbook(ruta)["Seguimiento"].max_row
        escribir(resultado, ruta)
        filas_tras_la_segunda = load_workbook(ruta)["Seguimiento"].max_row

        assert filas_tras_la_segunda == filas_tras_la_primera * 2 - 1


class TestMetricas:
    def test_el_resultado_cuadra_con_lo_leido(self, config):
        correos = leer_correos(CSV_CORREOS)
        resultado = ejecutar(correos, config)
        assert resultado.correos_leidos == len(correos)
        assert resultado.proveedor == "reglas"
        # Cada correo acaba en el seguimiento o en descartes; ninguno se pierde.
        remitentes_procesados = {f.id_correo for f in resultado.filas}
        remitentes_descartados = {d.correo.id for d in resultado.descartes}
        assert len(remitentes_procesados | remitentes_descartados) == len(correos)

    def test_el_porcentaje_automatizado_es_coherente(self, config):
        resultado = ejecutar(leer_correos(CSV_CORREOS), config)
        assert 0 <= resultado.porcentaje_automatizado <= 100
        assert len(resultado.automaticas) + len(resultado.para_revision) <= len(resultado.filas)
