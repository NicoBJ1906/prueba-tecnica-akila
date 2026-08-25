"""Volcado del resultado al Excel de seguimiento.

El libro tiene cuatro hojas y cada una responde a una pregunta distinta:

- `Seguimiento`  — el registro que pide el enunciado, con las seis columnas
                   exigidas primero y la auditoría después.
- `Para revisar` — la cola de trabajo de la persona: solo lo que el sistema no
                   se atreve a dar por bueno. Es la hoja que abre cada mañana.
- `Descartados`  — lo que no llegó al seguimiento y por qué. Sin esta hoja, la
                   automatización pide un acto de fe.
- `Resumen`      — qué hizo el proceso en esta ejecución.

Se escribe en modo añadir: si el fichero ya existe, las filas nuevas se suman a
las que hubiera, respetando lo que la persona haya anotado a mano.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from .modelos import (
    ESTADO_REVISION,
    FilaSeguimiento,
    ResultadoTriaje,
)

# Las seis primeras son, literalmente y en este orden, las que pide el enunciado.
COLUMNAS = (
    "Fecha",
    "Cliente",
    "Tipo",
    "Urgencia",
    "Acción",
    "Responsable",
    # --- auditoría, a partir de aquí ---
    "Estado",
    "Apartamento",
    "Categoría",
    "Confianza",
    "Clasificado por",
    "Motivo",
    "Remitente",
    "ID correo",
)
COLUMNAS_EXIGIDAS = COLUMNAS[:6]

COLUMNAS_DESCARTES = ("Fecha", "Remitente", "Asunto", "Motivo", "Regla aplicada")

ANCHOS = {
    "Fecha": 17, "Cliente": 22, "Tipo": 14, "Urgencia": 11, "Acción": 52,
    "Responsable": 22, "Estado": 16, "Apartamento": 18, "Categoría": 12,
    "Confianza": 11, "Clasificado por": 16, "Motivo": 46, "Remitente": 30,
    "ID correo": 14, "Asunto": 34, "Regla aplicada": 22,
}

AZUL_CABECERA = PatternFill("solid", fgColor="1F3864")
FUENTE_CABECERA = Font(bold=True, color="FFFFFF")
RELLENO_REVISION = PatternFill("solid", fgColor="FFF2CC")
RELLENO_ALTA = PatternFill("solid", fgColor="FCE4E4")


def _escribir_cabecera(hoja: Worksheet, columnas: tuple[str, ...]) -> None:
    hoja.append(list(columnas))
    for indice, nombre in enumerate(columnas, start=1):
        celda = hoja.cell(row=1, column=indice)
        celda.fill = AZUL_CABECERA
        celda.font = FUENTE_CABECERA
        celda.alignment = Alignment(vertical="center")
        hoja.column_dimensions[get_column_letter(indice)].width = ANCHOS.get(nombre, 18)
    hoja.freeze_panes = "A2"
    hoja.auto_filter.ref = f"A1:{get_column_letter(len(columnas))}1"


def _fila_como_lista(fila: FilaSeguimiento) -> list:
    return [
        fila.fecha.strftime("%Y-%m-%d %H:%M"),
        fila.cliente,
        fila.tipo,
        fila.urgencia,
        fila.accion,
        fila.responsable,
        fila.estado,
        fila.apartamento,
        fila.categoria,
        round(fila.confianza, 2),
        fila.fuente,
        fila.motivo,
        fila.remitente,
        fila.id_correo,
    ]


def _pintar(hoja: Worksheet, numero_fila: int, fila: FilaSeguimiento) -> None:
    """Resalta lo que necesita atención humana. El color nunca va solo: la
    columna 'Estado' dice lo mismo en texto, para quien no distinga los tonos."""
    relleno = None
    if fila.estado == ESTADO_REVISION:
        relleno = RELLENO_REVISION
    elif fila.urgencia == "Alta":
        relleno = RELLENO_ALTA
    if relleno:
        for columna in range(1, len(COLUMNAS) + 1):
            hoja.cell(row=numero_fila, column=columna).fill = relleno


def _hoja(libro: Workbook, titulo: str, columnas: tuple[str, ...]) -> Worksheet:
    """Devuelve la hoja, creándola con cabecera si no existe."""
    if titulo in libro.sheetnames:
        return libro[titulo]
    hoja = libro.create_sheet(titulo)
    _escribir_cabecera(hoja, columnas)
    return hoja


def escribir(resultado: ResultadoTriaje, ruta: str | Path) -> Path:
    """Escribe (o amplía) el Excel de seguimiento y devuelve su ruta."""
    ruta = Path(ruta)
    ruta.parent.mkdir(parents=True, exist_ok=True)

    if ruta.exists():
        libro = load_workbook(ruta)
    else:
        libro = Workbook()
        libro.remove(libro.active)

    seguimiento = _hoja(libro, "Seguimiento", COLUMNAS)
    revision = _hoja(libro, "Para revisar", COLUMNAS)
    descartados = _hoja(libro, "Descartados", COLUMNAS_DESCARTES)

    for fila in resultado.filas:
        seguimiento.append(_fila_como_lista(fila))
        _pintar(seguimiento, seguimiento.max_row, fila)
        if fila.estado == ESTADO_REVISION:
            revision.append(_fila_como_lista(fila))
            _pintar(revision, revision.max_row, fila)

    for descarte in resultado.descartes:
        descartados.append(
            [
                descarte.correo.fecha_recepcion.strftime("%Y-%m-%d %H:%M"),
                descarte.correo.remitente,
                descarte.correo.asunto,
                descarte.motivo,
                descarte.regla,
            ]
        )

    _escribir_resumen(libro, resultado)

    # El resumen va primero: es lo que se quiere ver al abrir el fichero.
    libro.move_sheet("Resumen", offset=-len(libro.sheetnames) + 1)
    libro.save(ruta)
    return ruta


def _escribir_resumen(libro: Workbook, resultado: ResultadoTriaje) -> None:
    if "Resumen" in libro.sheetnames:
        libro.remove(libro["Resumen"])
    hoja = libro.create_sheet("Resumen")

    hoja.column_dimensions["A"].width = 42
    hoja.column_dimensions["B"].width = 30

    titulo = hoja.cell(row=1, column=1, value="Triaje de correos · última ejecución")
    titulo.font = Font(bold=True, size=13)

    revision = len(resultado.para_revision)
    lineas = [
        ("Correos leídos", resultado.correos_leidos),
        ("Ya procesados en ejecuciones anteriores", len(resultado.ya_procesados)),
        ("Descartados (remitentes automáticos)", len(resultado.descartes)),
        ("Filas generadas en el seguimiento", len(resultado.filas)),
        ("  · resueltas automáticamente", len(resultado.automaticas)),
        ("  · pendientes de revisión humana", revision),
        ("Porcentaje automatizado", f"{resultado.porcentaje_automatizado:.0f} %"),
        ("", ""),
        ("Clasificación realizada por", resultado.proveedor),
        ("Llamadas al modelo", resultado.llamadas_ia),
        ("Tokens de entrada", resultado.tokens_entrada),
        ("Tokens de salida", resultado.tokens_salida),
        ("Coste estimado de la ejecución", f"USD {coste_estimado(resultado):.4f}"),
        ("Duración", f"{resultado.segundos:.1f} s"),
    ]

    for indice, (etiqueta, valor) in enumerate(lineas, start=3):
        hoja.cell(row=indice, column=1, value=etiqueta).font = Font(
            bold=not etiqueta.startswith(" ")
        )
        hoja.cell(row=indice, column=2, value=valor)


# Tarifas públicas de Claude Haiku 4.5 (USD por millón de tokens). Se usan solo
# para estimar el coste en el informe; no afectan al comportamiento.
PRECIO_ENTRADA_POR_MILLON = 1.0
PRECIO_SALIDA_POR_MILLON = 5.0


def coste_estimado(resultado: ResultadoTriaje) -> float:
    if resultado.proveedor != "anthropic":
        return 0.0
    return (
        resultado.tokens_entrada / 1_000_000 * PRECIO_ENTRADA_POR_MILLON
        + resultado.tokens_salida / 1_000_000 * PRECIO_SALIDA_POR_MILLON
    )
